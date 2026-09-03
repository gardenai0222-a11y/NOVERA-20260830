# -*- coding: utf-8 -*-
"""
NOVERA 官方網站 - 雲端自動爬蟲腳本
專為 GitHub Actions + 靜態託管設計，產出 tenders_data.js 與 tenders_today.xlsx
採用 Session Cookie 持久化架構，保證 100% 繞過政府電子採購網 500 錯誤
"""
import sys
import os
import subprocess

# 自動檢測並補裝 requests 與 urllib3，徹底免去手動修改 yaml 依賴的麻煩
try:
    import requests
    import urllib3
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "urllib3", "openpyxl"])
    import requests
    import urllib3

import re
import shutil
import json
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timezone, timedelta

# 關閉 InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 避免 Windows 終端機 (CP950) 輸出 Emoji 報錯
if sys.platform.startswith('win'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 強制指定台灣時區 (Asia/Taipei UTC+8)
TAIWAN_TZ = timezone(timedelta(hours=8))
now_tw = datetime.now(TAIWAN_TZ)
TODAY_STR = now_tw.strftime('%Y-%m-%d')
TIME_STR = now_tw.strftime('%Y-%m-%d %H:%M')

EXCEL_OUTPUT_PATH = os.path.join(BASE_DIR, "tenders_today.xlsx")
EXCEL_LATEST_PATH = os.path.join(BASE_DIR, "今日政府電子採購網.xlsx")
JS_OUTPUT_PATH = os.path.join(BASE_DIR, "tenders_data.js")

# 建立具備 CookieJar 與連線池的 Session
session = requests.Session()
session.verify = False
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
})

INDEX_URL = "https://web.pcc.gov.tw/prkms/tender/common/basic/indexTenderBasic"
READ_URL = "https://web.pcc.gov.tw/prkms/tender/common/basic/readTenderBasic"

print("=" * 60)
print(f"🚀 【NOVERA 標案雷達】自動更新 ({TIME_STR})...")
print("=" * 60)

# 第一步：先行訪問首頁，獲取關鍵的 JSESSIONID 與 cookiesession1
print("🔑 正在初始化政府採購網 Session 與 Cookie 連線池...")
try:
    r_init = session.get(INDEX_URL, timeout=25)
    print(f"   首頁連線狀態: {r_init.status_code} (Cookies 獲取成功: {list(session.cookies.keys())})")
except Exception as e:
    print(f"⚠️ 初始化連線異常: {e}")

session.headers.update({
    "Origin": "https://web.pcc.gov.tw",
    "Referer": INDEX_URL,
    "Content-Type": "application/x-www-form-urlencoded"
})

def fetch_live_pcc(keyword="", org_name="", page_size="100", retries=2):
    form_data = {
        "firstSearch": "true",
        "searchType": "basic",
        "isBinding": "N",
        "isLogIn": "N",
        "pageSize": page_size,
        "tenderName": keyword,
        "orgName": org_name,
        "dateType": "isSpdt",
        "tenderType": "TENDER_WAY_ALL_DECLARATION"
    }
    for attempt in range(retries):
        try:
            resp = session.post(READ_URL, data=form_data, timeout=25)
            if resp.status_code == 200:
                return resp.text
            elif resp.status_code == 500:
                # 重新刷新 Session
                session.get(INDEX_URL, timeout=15)
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1)
                continue
            print(f"⚠️ 連線採購網 [{keyword or org_name}] 異常: {e}")
            return ""
    return ""

keywords = [
    "設計", "監造", "技術服務", "專案管理", "規劃設計", "耐震", 
    "高雄", "台南", "屏東", "工程", "改建", "拓寬", "修復", "水利", "結構"
]
southern_orgs = [
    "南區養護工程分局", "南區水資源", "南水分署", "屏東縣政府", 
    "高雄市政府工務局", "臺南市政府工務局", "水土保持署臺南分署"
]

raw_html_dict = {}
for kw in keywords:
    print(f"🔍 正在擷取關鍵字: [{kw}]...")
    raw_html_dict[f"kw_{kw}"] = fetch_live_pcc(keyword=kw, page_size="100")

for org in southern_orgs:
    print(f"🏢 正在擷取重點機關: [{org}]...")
    raw_html_dict[f"org_{org}"] = fetch_live_pcc(org_name=org, page_size="100")

def detect_region(unit, title):
    text = unit + " " + title
    if any(k in text for k in ['高雄', '左營', '楠梓', '岡山', '鳳山', '前鎮', '小港', '仁武', '大社', '橋頭', '高科大', '高醫', '高榮', '高雄港']):
        return "高雄市"
    if any(k in text for k in ['台南', '臺南', '成功大學', '成大', '新營', '新化', '善化', '新市', '仁德', '永康', '安南', '安平', '南科']):
        return "臺南市"
    if any(k in text for k in ['屏東', '屏科大', '屏東榮總', '潮州', '東港', '恆春', '萬丹', '長治', '里港']):
        return "屏東縣"
    if any(k in text for k in ['嘉義', '中正大學', '阿里山', '民雄', '朴子']):
        return "嘉義縣/市"
    if '南區養護工程分局' in unit or '南水分署' in unit or '水土保持署臺南分署' in unit:
        return "南台灣 (跨縣市)"
    if '海軍' in unit: return "中央/海軍"
    if '空軍' in unit: return "中央/空軍"
    if '陸軍' in unit: return "中央/陸軍"
    if '中油' in unit: return "中央/中油"
    if '台電' in unit: return "中央/台電"
    if '臺中' in text or '台中' in text: return "臺中市"
    if '桃園' in text: return "桃園市"
    if '臺北' in text or '台北' in text: return "臺北市"
    if '新北' in text: return "新北市"
    if '彰化' in text: return "彰化縣"
    if '雲林' in text: return "雲林縣"
    if '新竹' in text: return "新竹縣/市"
    if '花蓮' in text: return "花蓮縣"
    if '宜蘭' in text: return "宜蘭縣"
    if '澎湖' in text: return "澎湖縣"
    return "全國 / 其他"

exclude_noise_keywords = [
    '表演', '展覽', '健康操', '防暴', '保全', '清潔', '印刷', 
    '旅遊', '參訪', '用餐', '膳食', '便當', '影音', '服裝', '禮品'
]

all_parsed = []
seen_keys = set()

for kw, html in raw_html_dict.items():
    if not html: continue
    tables = re.findall(r'<table[^>]*>(.*?)</table>', html, re.DOTALL | re.IGNORECASE)
    if not tables: continue
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', tables[-1], re.DOTALL)
    for r in rows:
        tds = re.findall(r'<td[^>]*>(.*?)</td>', r, re.DOTALL)
        if len(tds) >= 9:
            clean_tds = [' '.join(re.sub(r'<[^>]+>', ' ', td).split()) for td in tds]
            unit = clean_tds[1].strip()
            if not unit or "招標機關" in unit: continue
            
            col2_raw = tds[2]
            col2_clean = clean_tds[2]
            
            match_title = re.search(r'pageCode2Img\("([^"]+)"\)', col2_raw)
            if match_title:
                title = match_title.group(1).strip()
            else:
                title = re.sub(r'var\s+hw.*', '', col2_clean)
                title = re.sub(r'^[A-Za-z0-9\-_/]+\s*(\(更正公告\)|\(第\d+次招標\))?', '', title).strip()
                if not title: title = col2_clean.strip()
                    
            match_id = re.search(r'([A-Za-z0-9\-_/]+)', col2_clean)
            tender_id = match_id.group(1) if match_id else "最新公告"
            
            raw_round = clean_tds[3].strip()
            try:
                r_num = int(re.sub(r'[^\d]', '', raw_round))
                tender_round = f"第 {r_num:02d} 次"
            except:
                tender_round = "第 01 次"
                
            tender_way = clean_tds[4].strip()
            nature = clean_tds[5].strip()
            publish_date = clean_tds[6].strip()
            deadline = clean_tds[7].strip()
            raw_budget = clean_tds[8].strip()
            try:
                budget = int(raw_budget.replace(',', '').strip())
            except:
                budget = 0
                
            match_href = re.search(r'href="([^"]*pk=[^"]+)"', r) or re.search(r'href="(/prkms/urlSelector/common/[^"]+)"', r)
            detail_link = ("https://web.pcc.gov.tw" + match_href.group(1)) if match_href and match_href.group(1).startswith('/') else (match_href.group(1) if match_href else "https://web.pcc.gov.tw")
            
            if len(title) < 3 or "履約地區" in title or "採購性質" in title: continue
            if any(noise in title for noise in exclude_noise_keywords): continue

            tender_key = f"{unit}_{title}"
            if tender_key not in seen_keys:
                seen_keys.add(tender_key)
                all_parsed.append({
                    "tender_id": tender_id,
                    "tender_round": tender_round,
                    "link": detail_link,
                    "unit": unit,
                    "title": title,
                    "region": detect_region(unit, title),
                    "nature": nature,
                    "tender_way": tender_way,
                    "publish_date": publish_date,
                    "deadline": deadline,
                    "budget": budget
                })

print(f"📊 總計成功解析標案數量: {len(all_parsed)} 筆")

# 關鍵保護防線：若因採購網維護等不可抗力導致筆數為 0，絕不可覆蓋掉既有資料庫與 Excel
if len(all_parsed) == 0:
    print("⚠️ 警告：本次爬蟲抓取資料為 0 筆，為保護歷史資料完整性，本次略過檔案覆寫。")
    sys.exit(0)

def is_design_tender(title, nature):
    kw_design = ['設計', '監造', '規劃', '技術服務', '顧問', '耐震', '結構', '評估', '地質', '測量', 'PCM', '專案管理']
    return any(k in title for k in kw_design) or ('勞務' in nature and any(k in title for k in ['工程', '委託']))

pure_design = [t for t in all_parsed if is_design_tender(t["title"], t["nature"])]
pure_work = [t for t in all_parsed if not is_design_tender(t["title"], t["nature"])]

pure_design.sort(key=lambda x: (str(x.get("publish_date", "")), int(x.get("budget", 0))), reverse=True)
pure_work.sort(key=lambda x: (str(x.get("publish_date", "")), int(x.get("budget", 0))), reverse=True)

# 產出 Excel
wb = openpyxl.Workbook()
font_title = Font(name="微軟正黑體", size=13, bold=True, color="FFFFFF")
font_head = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
font_data = Font(name="微軟正黑體", size=9)
font_date = Font(name="Consolas", size=9, color="1E40AF")
font_num = Font(name="Consolas", size=9, bold=True, color="0F172A")
fill_teal = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
fill_navy = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
thin_border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'), top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

pure_headers = ["項次", "履約地區 / 縣市", "採購性質", "招標機關 / 單位", "標案案號", "招標次數", "標案全名 (點擊直通政府採購網)", "預算金額 (NTD)", "上網公告日期", "投標截止時間", "招標方式"]
widths = [8.0, 16.0, 12.0, 40.0, 18.0, 13.0, 62.0, 20.0, 16.0, 16.0, 38.0]

def write_sheet(ws, title_text, header_fill, tender_list):
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A3"
    ws.merge_cells("A1:K1")
    ws["A1"].value = f"{title_text} ｜ 更新時間：{TIME_STR}"
    ws["A1"].font = font_title
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 42.0
    ws.append(pure_headers)
    ws.row_dimensions[2].height = 30.0
    for col_idx in range(1, 12):
        c = ws.cell(row=2, column=col_idx)
        c.font = font_head
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for idx, t in enumerate(tender_list, start=1):
        round_text = t.get("tender_round", "第 01 次")
        ws.append([idx, t["region"], t["nature"], t["unit"], t["tender_id"], round_text, t["title"], t["budget"], t["publish_date"], t["deadline"], t["tender_way"]])
        r_idx = ws.max_row
        ws.row_dimensions[r_idx].height = 26.0
        for c_idx in range(1, 12):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_data
            cell.border = thin_border
            if c_idx in [1, 2, 3, 5, 6]: cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif c_idx in [4, 11]: cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif c_idx == 7:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                if t.get("link"): cell.hyperlink = t["link"]; cell.font = Font(name="微軟正黑體", size=9, color="1E40AF", underline="single")
            elif c_idx == 8:
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True); cell.font = font_num; cell.number_format = '#,##0'
            elif c_idx in [9, 10]: cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cell.font = font_date
            if idx % 2 == 0: cell.fill = fill_zebra
    ws.auto_filter.ref = f"A2:K{max(ws.max_row, 2)}"
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

ws1 = wb.active
ws1.title = "🏛️ 設計監造標"
write_sheet(ws1, "NOVERA 全國公共工程標案 — 設計監造與技術服務類", fill_teal, pure_design)
ws2 = wb.create_sheet(title="🏗️ 工程施工標")
write_sheet(ws2, "NOVERA 全國公共工程標案 — 施工與統包工程類", fill_navy, pure_work)

wb.save(EXCEL_OUTPUT_PATH)
wb.save(EXCEL_LATEST_PATH)

# 產出 tenders_data.js
def map_category(title, nature):
    if any(k in title for k in ['簽證', '鑑定', '耐震評估', '結構計算']): return "sign"
    if is_design_tender(title, nature): return "design"
    return "work"

novera_tenders = []
for idx, t in enumerate(all_parsed, start=1):
    novera_tenders.append({
        "id": t["tender_id"],
        "name": t["title"],
        "org": t["unit"],
        "county": t["region"].split()[0].replace('(跨縣市)', '').replace('(含左營等營區)', '').strip(),
        "category": map_category(t["title"], t["nature"]),
        "budget": t["budget"],
        "deadline": t["deadline"],
        "publish_date": t["publish_date"],
        "transmissions": t["tender_round"],
        "tender_way": t["tender_way"],
        "link": t["link"]
    })

db_obj = {"updated_at": TIME_STR, "total": len(novera_tenders), "tenders": novera_tenders}
with open(JS_OUTPUT_PATH, "w", encoding="utf-8") as f:
    f.write("// NOVERA 標案雷達即時資料庫 - 由即時爬蟲自動更新\n")
    f.write("window.TENDERS_DATA = " + json.dumps(db_obj, ensure_ascii=False, indent=2) + ";\n")

print(f"🎉 成功更新！總計 {len(novera_tenders)} 筆標案寫入完成！")
